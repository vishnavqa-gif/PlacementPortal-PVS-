from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from datetime import date, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from app.database.session import get_db
from app.models.models import User, Student, Interview
from app.core.security import require_admin

router = APIRouter(prefix="/reports", tags=["Reports"])


async def fetch_interview_data(db: AsyncSession, start_date: date, end_date: date):
    result = await db.execute(
        select(Interview).where(
            Interview.interview_date >= start_date,
            Interview.interview_date <= end_date,
        )
    )
    return result.scalars().all()


@router.get("/daily")
async def daily_report(
    report_date: date = Query(default=date.today()),
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    interviews = await fetch_interview_data(db, report_date, report_date)
    
    statuses = {}
    for i in interviews:
        statuses[i.status] = statuses.get(i.status, 0) + 1

    return {
        "date": str(report_date),
        "total": len(interviews),
        "by_status": statuses,
    }


@router.get("/weekly")
async def weekly_report(
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    end_date = date.today()
    start_date = end_date - timedelta(days=7)
    interviews = await fetch_interview_data(db, start_date, end_date)

    daily = {}
    for i in interviews:
        d = str(i.interview_date)
        daily[d] = daily.get(d, 0) + 1

    return {
        "start_date": str(start_date),
        "end_date": str(end_date),
        "total": len(interviews),
        "daily_breakdown": daily,
    }


@router.get("/monthly")
async def monthly_report(
    year: int = Query(default=date.today().year),
    month: int = Query(default=date.today().month),
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    from calendar import monthrange
    _, last_day = monthrange(year, month)
    start_date = date(year, month, 1)
    end_date = date(year, month, last_day)
    interviews = await fetch_interview_data(db, start_date, end_date)

    return {
        "year": year,
        "month": month,
        "total": len(interviews),
        "completed": sum(1 for i in interviews if i.status == "completed"),
        "selected": sum(1 for i in interviews if i.status == "selected"),
        "rejected": sum(1 for i in interviews if i.status == "rejected"),
        "cancelled": sum(1 for i in interviews if i.status == "cancelled"),
    }


@router.get("/export/excel")
async def export_excel(
    start_date: date = Query(default=date.today() - timedelta(days=30)),
    end_date: date = Query(default=date.today()),
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    interviews = await fetch_interview_data(db, start_date, end_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Interviews Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")

    headers = ["#", "Interview Date", "Status", "Round", "Technology", "Interviewer", "Cabin", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, i in enumerate(interviews, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=str(i.interview_date))
        ws.cell(row=row, column=3, value=i.status)
        ws.cell(row=row, column=4, value=i.current_round)
        ws.cell(row=row, column=5, value=str(i.technology_id))
        ws.cell(row=row, column=6, value=str(i.interviewer_id))
        ws.cell(row=row, column=7, value=str(i.cabin_id))
        ws.cell(row=row, column=8, value=i.notes or "")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=interviews_{start_date}_{end_date}.xlsx"}
    )


@router.get("/export/pdf")
async def export_pdf(
    start_date: date = Query(default=date.today() - timedelta(days=30)),
    end_date: date = Query(default=date.today()),
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    interviews = await fetch_interview_data(db, start_date, end_date)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Interview Report", styles["Title"]))
    elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    data = [["#", "Date", "Status", "Round"]]
    for idx, i in enumerate(interviews, 1):
        data.append([str(idx), str(i.interview_date), i.status, i.current_round])

    table = Table(data, colWidths=[30, 100, 100, 120])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1976D2")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=interviews_{start_date}_{end_date}.pdf"}
    )
